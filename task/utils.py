from errno import EISDIR
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from string import ascii_uppercase
from config.settings import BASE_DIR
import os
import json

def closet_base(width, depth, height, num_shelves, num_vertical_bar=None):
    """
    선반 default : 800 x 400, 15,000원. width, depth에 선형 비례
    세로 벽 default : 1900 x 400, 45,000원. depth에 선형 비례, 높이 2000 초과 시 60,000
    뒷면 default 800 x 1900, 24,000원. 너비에 선형 비례
    문짝 400 x 1900, 30,000. 넓이에 선형 비례

    :param width: 장의 너비
    :param depth: 장의 깊이
    :param height: 장의 높이
    :param num_vertical_bar: 장의 세로 축 개수
    :param num_shelves: 장의 선반 수
    :param num_doors: 문짝 수
    :return: 가격
    """
    unit_price_dict = {
        'shelves': 15000,
        'walls': 60000 if height > 2000 else 45000
    }

    if num_vertical_bar < 0:
        num_vertical_bar = width // 900

    shelves = (num_shelves + 1) * (width / 800) * (depth / 400) * unit_price_dict['shelves']
    if width > 1200:
        shelves *= 2
    walls = (num_vertical_bar + 2) * (depth / 400) * unit_price_dict['walls']

    return [shelves, walls]


def calc_closet(estimator):
    """
    선반 default : 800 x 400, 15,000원. width, depth에 선형 비례
    세로 벽 default : 1900 x 400, 45,000원. depth에 선형 비례, 높이 2000 초과 시 60,000
    뒷면 default 800 x 1900, 24,000원. 너비에 선형 비례
    문짝 400 x 1900, 30,000. 넓이에 선형 비례

    :return: 가격
    """

    width, depth, height, num_shelves, num_vertical_bar, num_doors  = int(estimator['width']), int(estimator['depth']), int(estimator['height']), int(estimator['num_shelves']), int(estimator['num_vertical_bar']), int(estimator['num_doors'])


    unit_price_dict = {
        'shelves': 15000,
        'walls': 60000 if height > 2000 else 45000,
        'back': 24000,
        'doors': 45000
    }

    # width가 1200 미만일 때
    # 1200 넘길 때 가격 변동은 선반만 고려하면 될ㄷ스
    shelves, walls = closet_base(width, depth, height, num_shelves, num_vertical_bar=num_vertical_bar)
    back = (width / 800) * unit_price_dict['back']

    if num_doors > 0:
        doors = unit_price_dict['doors'] * num_doors if width / num_doors < 400 else (width / (400 * num_doors)) * unit_price_dict['doors'] * num_doors
    else:
        doors = 0

    total = shelves + walls + back + doors

    return {'shelves' : shelves, 'walls' : walls, 'back' : back, 'doors' : doors, 'total' : total}

def rgb2hex(rgb):
    return ''.join([ hex(c)[2:].zfill(2) for c in rgb])



def render_estimator_base(sheet, estimator, container):
    """
    sheet, estimator를 받아서 값을 채우기

    """

    border1 = Border(bottom=Side(border_style='medium', color=rgb2hex([176, 176, 176])))
    border2 = Border(bottom=Side(border_style='medium', color=rgb2hex([120, 120, 120])))
    border3 = Border(bottom=Side(border_style='thick', color=rgb2hex([255, 192, 0])))

    # estimator information
    sheet['B9'] = container.id
    sheet['B12'] = container.author.name
    sheet['B15'] = container.create_date.strftime('%Y-%m-%d %H:%M:%S')


    # 각 항목별 값 채우기
    i = 0
    for k, v in estimator.prices.items():
        if k == 'total':
            continue
        sheet[f'E{19 + i * 3}'].value = k
        sheet[f'J{19 + i * 3}'].value = v
        # load_ws[f'L{14 + i * 3}'].value = 수량인데 일단은 없자낭..
        sheet[f'M{19 + i * 3}'].value = v

        # 하단 테두리
        for j in range(4, 15):
            sheet[f'{ascii_uppercase[j]}{20 + i * 3}'].border = border1

        i += 1

    if estimator.additional_kwargs:
        for k, v in estimator.additional_kwargs.items():
            sheet[f'E{19 + i * 3}'].value = k
            sheet[f'J{19 + i * 3}'].value = v
            # load_ws[f'L{14 + i * 3}'].value = 수량인데 일단은 없자낭..
            sheet[f'M{19 + i * 3}'].value = v

            # 하단 테두리
            for j in range(4, 15):
                sheet[f'{ascii_uppercase[j]}{20 + i * 3}'].border = border1

            i += 1


    # 하단 테두리
    i -= 1
    for j in range(4, 15):
        sheet[f'{ascii_uppercase[j]}{20 + i * 3}'].border = border2
    # 다음 세칸
    i += 1

    # 하단 테두리
    for j in range(4, 15):
        sheet[f'{ascii_uppercase[j]}{21 + i * 3}'].border = border3

    # 합계
    sheet[f'E{19 + i * 3}'].value = '합계'
    sheet[f'E{20 + i * 3}'].value = 'VAT'
    sheet[f'E{24 + i * 3}'].value = '총 금액'

    # 수식할당
    sheet[f'M{19 + i * 3}'].value = f'=SUM(M14:M{16 + i * 3})'
    sheet[f'M{20 + i * 3}'].value = f'=M{19 + i * 3} / 10'
    sheet[f'M{24 + i * 3}'].value = f'=M{19 + i * 3} + M{20 + i * 3}'



def render_esitmator(estimator):
    file_path = os.path.join(BASE_DIR, 'data/invoice_template.xlsx')
    template = load_workbook(file_path)

    sheet = template['Sheet1']

    render_estimator_base(sheet, estimator)
    # 저장
    file_path = os.path.join(BASE_DIR, f'data/{estimator.name}.xlsx')
    img_path = os.path.join(BASE_DIR, f'data/esitmation.png')

    template.save(file_path)


def config_parser(f_name):

    path = f'task/configs/{f_name}.json'
    with open(path, 'r+', encoding='utf-8') as f:
        config = json.load(f)

    name = config['name']
    del config['name']

    return config, name


def render_container(container):
    file_path = os.path.join(BASE_DIR, 'data/invoice_template.xlsx')
    template = load_workbook(file_path)

    sheet_origin = template['Sheet1']

    

    for i, estimator in enumerate(container.estimator_set.all()):
        sheet = template.copy_worksheet(sheet_origin)
        sheet.title = f'title {i}'
        render_estimator_base(sheet, estimator, container)

    

    # 저장
    file_path = os.path.join(BASE_DIR, f'data/{container.name}.xlsx')
    img_path = os.path.join(BASE_DIR, f'data/esitmation.png')

    template.save(file_path)
